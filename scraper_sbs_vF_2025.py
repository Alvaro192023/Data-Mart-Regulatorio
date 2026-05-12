"""
============================================================================
SCRAPER SBS vF_2025 - Directorio de Empresas Supervisadas
============================================================================
Extrae datos del directorio de empresas supervisadas por la SBS
(Superintendencia de Banca, Seguros y AFP del Peru).

NO hay PDFs descargables en el portal de la SBS (a diferencia de la SMV).
Los dictamenes y EEFF auditados los publica cada empresa en su propio
sitio web. Este scraper extrae solo datos del directorio.

Categorias:
  - Empresas Bancarias, Financieras, Cajas Municipales, Cajas Rurales
  - Edpymes, Empresas de Seguros, AFPs, Empresas de Creditos
  - Bancos de Inversion, Afianzadoras, Servicios Fiduciarios

Datos extraidos (del portal SBS):
  - Razon Social, RUC, Tipo de Empresa, Direccion
  - Representante Legal, Telefono, Sitio Web

Columnas vacias (para Power Automate futuro):
  - Auditor, Fee Auditoria, Tipo de Opinion, Directorio

Flujo:
  1. Ejecutar este script -> genera Excel con directorio completo
  2. Cruzar con data SMV para identificar overlap
  3. (Futuro) Power Automate para extraer auditor/fee de PDFs individuales

Requisitos:
    pip install selenium openpyxl

Uso:
    python scraper_sbs_vF_2025.py                    # Completo
    python scraper_sbs_vF_2025.py --max 1            # Solo 1 categoria (prueba)
    python scraper_sbs_vF_2025.py --no-headless      # Ver navegador
============================================================================
"""

import time, re, sys, argparse, logging
from datetime import datetime
from pathlib import Path

import requests as req

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================================
# CONFIGURACION
# ============================================================================
BASE_URL = "https://www.sbs.gob.pe"

# URLs directas de los iframes que contienen las tablas de datos
# Descubiertas del HTML: /app/sadel/Paginas/Redir/ListarFuncionarios.aspx?codTpEntidad=XX
IFRAME_BASE = f"{BASE_URL}/app/sadel/Paginas/Redir/ListarFuncionarios.aspx?codTpEntidad="
# (Nombre, Tipo Empresa, URL, Categoria)
CATEGORIAS = [
    # Directorio del Sistema Financiero
    ("Empresas Bancarias",          "Banco",                 f"{IFRAME_BASE}B",  "Sistema Financiero"),
    ("Empresas Financieras",        "Financiera",            f"{IFRAME_BASE}F",  "Sistema Financiero"),
    ("Cajas Municipales",           "Caja Municipal",        f"{IFRAME_BASE}C",  "Sistema Financiero"),
    ("Cajas Rurales",               "Caja Rural",            f"{IFRAME_BASE}R",  "Sistema Financiero"),
    ("Edpymes",                     "Edpyme",                f"{IFRAME_BASE}EP", "Sistema Financiero"),
    ("Empresas de Creditos",        "Empresa de Credito",    f"{IFRAME_BASE}E",  "Sistema Financiero"),
    ("Bancos de Inversion",         "Banco de Inversion",    f"{IFRAME_BASE}BI", "Sistema Financiero"),
    ("Afianzadoras y Garantias",    "Afianzadora",           f"{IFRAME_BASE}G",  "Sistema Financiero"),
    ("Servicios Fiduciarios",       "Servicios Fiduciarios", f"{IFRAME_BASE}FD", "Sistema Financiero"),
    ("Arrendamiento Financiero",    "Arrendamiento Financiero",f"{IFRAME_BASE}AF","Sistema Financiero"),
    ("Administradoras Hipotecarias","Adm. Hipotecaria",      f"{IFRAME_BASE}AH", "Sistema Financiero"),
    ("Fondo Mivivienda",            "Fondo Mivivienda",      f"{IFRAME_BASE}FO", "Sistema Financiero"),
    # Directorio del Sistema de Seguros
    ("Empresas de Seguros",         "Seguros",               f"{IFRAME_BASE}S",  "Sistema de Seguros"),
    # Directorio del Sistema Privado de Pensiones
    ("AFPs",                        "AFP",                   f"{IFRAME_BASE}A",  "Sistema Privado de Pensiones"),
    # Directorio de Empresas de Servicios Complementarios y Conexos
    ("Transferencia de Fondos",     "Transferencia Fondos",  f"{IFRAME_BASE}TF", "Servicios Complementarios"),
    ("Almacenes Generales Deposito","Almacen Deposito",      f"{IFRAME_BASE}AG", "Servicios Complementarios"),
    ("Transporte y Custodia Numerario","Transporte Numerario",f"{IFRAME_BASE}TC","Servicios Complementarios"),
    ("Emisoras Dinero Electronico", "Dinero Electronico",    f"{IFRAME_BASE}ED", "Servicios Complementarios"),
    # Directorio de Otras Empresas Supervisadas
    ("Cajas y Derramas",            "Caja/Derrama",          f"{IFRAME_BASE}DE", "Otras Supervisadas"),
    ("Fondo de Cajas Municipales",  "Fondo Cajas",           f"{IFRAME_BASE}FF", "Otras Supervisadas"),
    ("Empresas de Factoring",       "Factoring",             f"{IFRAME_BASE}FA", "Otras Supervisadas"),
]

# PDFs descargables (no tienen iframe)
COOPAC_URL = f"{BASE_URL}/supervisados-y-registros/registros/coopac-registradas"
AFOCAT_URL = f"{BASE_URL}/regulacion/afocat"

DELAY = (1.5, 3.0)
EXCEL_NAME = "sbs_directorio.xlsx"

# Ruta output (cambiar a SharePoint cuando este listo)
#SHAREPOINT_DIR = Path(r"C:\Users\avillanuev044\OneDrive - PwC\PE-IFS-SALES - Entidades de Regulación\SMV")
SHAREPOINT_DIR = Path("output")
OUTPUT_DIR = Path("output")
PDF_DIR = Path("output") / "pdfs_sbs"
LOG_DIR = Path("output")
TS = datetime.now().strftime("%Y%m%d_%H%M%S")

# ============================================================================
# LOGGING
# ============================================================================
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
PDF_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_DIR / f"scraper_sbs_{TS}.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


# ============================================================================
# UTILIDADES
# ============================================================================
def extraer_tipo_sociedad(nombre: str) -> str:
    """Extrae el tipo de sociedad del nombre de la empresa."""
    patrones = [
        (r'\bS\.?\s*A\.?\s*A\.?\b', 'S.A.A.'),
        (r'\bS\.?\s*A\.?\s*C\.?\b', 'S.A.C.'),
        (r'\bS\.?\s*A\.?\b(?!\.\s*[AC])', 'S.A.'),
        (r'\bS\.?\s*Civil\b', 'S. Civil'),
        (r'\bS\.?\s*C\.?\s*R\.?\s*L\.?\b', 'S.C.R.L.'),
        (r'\bS\.?\s*R\.?\s*L\.?\b', 'S.R.L.'),
        (r'\bSUCURSAL\b', 'Sucursal'),
    ]
    for patron, tipo in patrones:
        if re.search(patron, nombre, re.IGNORECASE):
            return tipo
    return ""


def crear_driver(headless=True):
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    try:
    #Operador Google
        from webdriver_manager.chrome import ChromeDriverManager
        svc = ChromeService(ChromeDriverManager().install())
    except ImportError:
        svc = ChromeService()
    #svc = ChromeService(r"C:\Users\avillanuev044\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe")
    d = webdriver.Chrome(service=svc, options=opts)
    d.set_page_load_timeout(45)
    d.implicitly_wait(10)
    return d


# ============================================================================
# SCRAPING PRINCIPAL
# ============================================================================
def scrape_categoria(driver, nombre_cat, tipo, url, categoria):
    """Scrapea una categoria del directorio SBS."""
    log.info(f"\n  Categoria: {nombre_cat}")
    log.info(f"  URL: {url}")

    empresas = []
    max_intentos = 3

    for intento in range(1, max_intentos + 1):
        try:
            driver.get(url)
            time.sleep(5)  # Esperar JS rendering

            # Buscar tablas en la pagina
            tablas = driver.find_elements(By.TAG_NAME, "table")
            if not tablas:
                # Intentar esperar mas
                time.sleep(5)
                tablas = driver.find_elements(By.TAG_NAME, "table")

            if tablas:
                log.info(f"  Tablas encontradas: {len(tablas)}")
                for tabla in tablas:
                    empresas_tabla = extraer_de_tabla(tabla, tipo, categoria)
                    empresas.extend(empresas_tabla)
                if empresas:
                    break

            # Si no hay tablas, buscar bloques de datos
            if not empresas:
                bloques = driver.find_elements(By.CSS_SELECTOR,
                    ".panel-body, .card-body, .accordion-body, details, "
                    "[class*='empresa'], [class*='entidad'], [class*='directorio']"
                )
                if bloques:
                    log.info(f"  Bloques encontrados: {len(bloques)}")
                    for bloque in bloques:
                        emp = extraer_de_bloque(bloque, tipo, categoria)
                        if emp:
                            empresas.append(emp)
                    if empresas:
                        break

            # Si aun nada, guardar HTML para debug
            if not empresas:
                debug_file = LOG_DIR / f"debug_sbs_{tipo.lower().replace(' ', '_')}.html"
                debug_file.write_text(driver.page_source, encoding="utf-8")
                page_text = driver.find_element(By.TAG_NAME, "body").text
                log.info(f"  Sin datos estructurados. Texto pagina ({len(page_text)} chars):")
                log.info(f"  {page_text[:300]}...")
                log.info(f"  HTML guardado: {debug_file}")

                if intento < max_intentos:
                    log.warning(f"  Reintentando ({intento}/{max_intentos})...")
                    time.sleep(5)
                    continue
                break

        except TimeoutException:
            log.error(f"  Timeout cargando {url}")
            if intento < max_intentos:
                time.sleep(10)
                continue
        except Exception as e:
            error_msg = str(e)
            if "ERR_NAME_NOT_RESOLVED" in error_msg or "net::" in error_msg:
                if intento < max_intentos:
                    log.warning(f"  Error de red (intento {intento}/{max_intentos}). Reintentando en 10s...")
                    time.sleep(10)
                    continue
                else:
                    log.error(f"  Error de red persistente: {error_msg[:80]}")
            else:
                log.error(f"  Error: {error_msg[:120]}")
            break

    log.info(f"  Empresas extraidas: {len(empresas)}")
    return empresas


def extraer_de_tabla(tabla, tipo, categoria):
    """Extrae datos de una tabla HTML."""
    empresas = []
    filas = tabla.find_elements(By.TAG_NAME, "tr")
    if len(filas) < 2:
        return empresas

    # Detectar headers
    headers = []
    for th in filas[0].find_elements(By.CSS_SELECTOR, "th, td"):
        h = re.sub(r'\s+', ' ', th.text.strip()).upper()
        headers.append(h)

    if not headers or len(headers) < 2:
        return empresas

    # Mapear columnas por nombre
    col = {}
    for idx, h in enumerate(headers):
        hl = h.lower()
        if any(k in hl for k in ["razón", "razon", "empresa", "nombre", "entidad", "denominación"]):
            col["razon_social"] = idx
        elif "ruc" in hl:
            col["ruc"] = idx
        elif any(k in hl for k in ["direcc", "domicilio"]):
            col["direccion"] = idx
        elif any(k in hl for k in ["representante", "gerente", "funcionario"]):
            col["representante"] = idx
        elif "cargo" in hl:
            col["cargo"] = idx
        elif any(k in hl for k in ["teléf", "telef", "fono", "central"]):
            col["telefono"] = idx
        elif "fax" in hl:
            col["fax"] = idx
        elif any(k in hl for k in ["web", "página", "pagina", "portal", "url", "sitio"]):
            col["web"] = idx

    log.info(f"    Headers: {headers}")
    log.info(f"    Mapeo: {col}")

    # Extraer filas de datos
    for fila in filas[1:]:
        celdas = fila.find_elements(By.TAG_NAME, "td")
        if not celdas or len(celdas) < 2:
            continue

        def get_celda(key):
            if key in col and col[key] < len(celdas):
                return re.sub(r'\s+', ' ', celdas[col[key]].text.strip())
            return ""

        razon = get_celda("razon_social")
        if not razon or len(razon) < 3:
            razon = re.sub(r'\s+', ' ', celdas[0].text.strip())
        if not razon or len(razon) < 3:
            continue

        # Representante: combinar cargo + funcionario si existen
        representante = get_celda("representante")
        cargo = get_celda("cargo")
        if representante and cargo:
            representante = f"{representante} ({cargo})"
        elif not representante and cargo:
            representante = cargo

        # Buscar RUC en toda la fila si no hay columna
        ruc = get_celda("ruc")
        if not ruc:
            texto_fila = fila.text
            ruc_match = re.search(r'\b(20\d{9})\b', texto_fila)
            ruc = ruc_match.group(1) if ruc_match else ""

        # Telefono + Fax
        telefono = get_celda("telefono")
        fax = get_celda("fax")
        if fax and fax != telefono:
            telefono = f"{telefono} / Fax: {fax}" if telefono else f"Fax: {fax}"

        # Buscar web en links de la fila
        web = get_celda("web")
        if not web:
            for link in fila.find_elements(By.TAG_NAME, "a"):
                href = link.get_attribute("href") or ""
                if href and "sbs.gob.pe" not in href and href.startswith("http"):
                    web = href
                    break

        empresas.append({
            "razon_social": razon,
            "ruc": ruc,
            "tipo_sociedad": extraer_tipo_sociedad(razon),
            "categoria": categoria,
            "tipo_empresa": tipo,
            "direccion": get_celda("direccion"),
            "representante": representante,
            "telefono": telefono,
            "web": web,
            "auditor": "",
            "fee": "",
            "tipo_opinion": "",
            "directorio_miembros": "",
        })

    return empresas


def extraer_de_bloque(bloque, tipo, categoria):
    """Extrae datos de un bloque/div individual."""
    texto = re.sub(r'\s+', ' ', bloque.text.strip())
    if not texto or len(texto) < 10:
        return None

    # Razon social: primer texto en negrita o primera linea
    razon = ""
    try:
        bold = bloque.find_elements(By.CSS_SELECTOR, "strong, b, h3, h4, .titulo, .nombre")
        if bold:
            razon = re.sub(r'\s+', ' ', bold[0].text.strip())
    except Exception:
        pass
    if not razon:
        lineas = texto.split('\n')
        razon = lineas[0].strip() if lineas else ""
    if not razon or len(razon) < 3:
        return None

    # RUC
    ruc_match = re.search(r'\b(20\d{9})\b', texto)
    ruc = ruc_match.group(1) if ruc_match else ""

    # Direccion
    dir_match = re.search(r'(?:Direcci[oó]n|Domicilio|Dir)[:\s]*(.*?)(?:Tel[eéf]|Fono|RUC|Representante|$)', texto, re.IGNORECASE)
    direccion = dir_match.group(1).strip() if dir_match else ""

    # Representante
    rep_match = re.search(r'(?:Representante\s*Legal|Gerente\s*General|Gerente)[:\s]*(.*?)(?:Tel[eéf]|Direcci|RUC|$)', texto, re.IGNORECASE)
    representante = rep_match.group(1).strip() if rep_match else ""

    # Telefono
    tel_match = re.search(r'(?:Tel[eéf]fono|Telf|Tel|Fono|Central)[:\s]*([0-9()\s\-+]+)', texto, re.IGNORECASE)
    telefono = tel_match.group(1).strip() if tel_match else ""

    # Web
    web = ""
    try:
        for link in bloque.find_elements(By.TAG_NAME, "a"):
            href = link.get_attribute("href") or ""
            if href and "sbs.gob.pe" not in href and href.startswith("http"):
                web = href
                break
    except Exception:
        pass

    return {
        "razon_social": razon,
        "ruc": ruc,
        "tipo_sociedad": extraer_tipo_sociedad(razon),
        "categoria": categoria,
        "tipo_empresa": tipo,
        "direccion": direccion,
        "representante": representante,
        "telefono": telefono,
        "web": web,
        "auditor": "",
        "fee": "",
        "tipo_opinion": "",
        "directorio_miembros": "",
    }


def deduplicar_empresas(empresas):
    """Consolida multiples filas por empresa en un solo registro.
    La tabla SBS tiene una fila por funcionario — una empresa puede tener
    Gerente General, Presidente del Directorio, etc. en filas separadas.
    Combina todos los representantes con ; separador.
    """
    grupos = {}
    for emp in empresas:
        key = (emp["razon_social"].upper().strip(), emp["tipo_empresa"])
        if key not in grupos:
            grupos[key] = emp.copy()
            grupos[key]["_todos_representantes"] = [emp["representante"]]
        else:
            existente = grupos[key]
            # Agregar representante si no es duplicado
            if emp["representante"] and emp["representante"] not in existente["_todos_representantes"]:
                existente["_todos_representantes"].append(emp["representante"])
            # Complementar datos vacios
            if not existente["direccion"] and emp["direccion"]:
                existente["direccion"] = emp["direccion"]
            if not existente["telefono"] and emp["telefono"]:
                existente["telefono"] = emp["telefono"]
            if not existente["web"] and emp["web"]:
                existente["web"] = emp["web"]
            if not existente["ruc"] and emp["ruc"]:
                existente["ruc"] = emp["ruc"]

    # Combinar representantes y limpiar
    resultado = []
    for emp in grupos.values():
        emp["representante"] = "; ".join(emp["_todos_representantes"])
        del emp["_todos_representantes"]
        resultado.append(emp)

    return resultado


def descargar_pdf_sbs(driver, url_pagina, buscar_textos, nombre_archivo, match_index=0):
    """Descarga un PDF del portal SBS usando cookies de Selenium.
    buscar_textos: lista de textos a buscar en links (en orden de prioridad)
    match_index: cual coincidencia tomar (0=primera, 1=segunda, etc.)
    """
    log.info(f"\n  Descargando: {nombre_archivo}")
    log.info(f"  URL: {url_pagina}")
    try:
        driver.get(url_pagina)
        time.sleep(5)

        # Buscar links SOLO dentro del contenido principal (no menu/footer)
        contenido_links = driver.find_elements(By.CSS_SELECTOR,
            "#dnn_ContentPane a, .DNNModuleContent a, .pagina-contenido a")
        # Fallback a todos los links si no hay contenido
        if not contenido_links:
            contenido_links = driver.find_elements(By.TAG_NAME, "a")

        pdf_url = None

        # Buscar por texto (sin exigir .pdf en href)
        for buscar in buscar_textos:
            matches = []
            for link in contenido_links:
                href = link.get_attribute("href") or ""
                text = link.text.strip().lower()
                if buscar.lower() in text and href:
                    matches.append(href)
            if matches and match_index < len(matches):
                pdf_url = matches[match_index]
                log.info(f"  Link encontrado (match {match_index}): '{buscar}' → {pdf_url}")
                break

        if not pdf_url:
            log.warning(f"  No se encontro link al PDF")
            return

        if not pdf_url.startswith("http"):
            pdf_url = BASE_URL + pdf_url
        log.info(f"  PDF URL: {pdf_url}")

        # Transferir cookies de Selenium a requests para evitar 403
        session = req.Session()
        for cookie in driver.get_cookies():
            session.cookies.set(cookie['name'], cookie['value'])
        session.headers.update({
            "User-Agent": driver.execute_script("return navigator.userAgent"),
            "Referer": url_pagina,
        })

        r = session.get(pdf_url, timeout=30, allow_redirects=True)
        if r.status_code == 200 and len(r.content) > 1024:
            pdf_path = PDF_DIR / nombre_archivo
            pdf_path.write_bytes(r.content)
            log.info(f"  Guardado: {pdf_path} ({len(r.content)//1024} KB)")
        else:
            log.warning(f"  Error HTTP {r.status_code} o archivo muy pequeno ({len(r.content)} bytes)")
        session.close()

    except Exception as e:
        log.error(f"  Error descargando {nombre_archivo}: {str(e)[:100]}")


def scrape(driver, max_cat=None):
    """Scrapea todas las categorias del directorio SBS."""
    categorias = CATEGORIAS[:max_cat] if max_cat else CATEGORIAS
    log.info(f"Categorias a scrapear: {len(categorias)} + COOPAC + AFOCAT (PDFs)")

    resultados = []
    for nombre_cat, tipo, url, categoria in categorias:
        empresas = scrape_categoria(driver, nombre_cat, tipo, url, categoria)
        antes = len(empresas)
        empresas = deduplicar_empresas(empresas)
        if antes != len(empresas):
            log.info(f"  Deduplicado: {antes} filas → {len(empresas)} empresas unicas")
        resultados.extend(empresas)
        time.sleep(2)

    # Descargar PDFs al final
    descargar_pdf_sbs(driver, COOPAC_URL,
        ["ver coopac"], "COOPAC_Registradas.pdf")
    descargar_pdf_sbs(driver, AFOCAT_URL,
        ["descargar archivo"], "AFOCAT_Registro.pdf", match_index=0)

    return resultados


# ============================================================================
# EXPORTAR EXCEL
# ============================================================================
def exportar(datos, filepath):
    wb = Workbook()
    HF = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HFL = PatternFill("solid", fgColor="1F4E79")
    HA = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CF = Font(name="Arial", size=10)
    CA = Alignment(vertical="top", wrap_text=True)
    BRD = Border(
        left=Side("thin", "B0B0B0"), right=Side("thin", "B0B0B0"),
        top=Side("thin", "B0B0B0"), bottom=Side("thin", "B0B0B0"),
    )
    ALT = PatternFill("solid", fgColor="E8F0FE")

    # --- Hoja principal ---
    ws = wb.active
    ws.title = "Directorio_SBS"
    headers = [
        ("Razon Social", 55),
        ("RUC", 16),
        ("Tipo Sociedad", 14),
        ("Categoria", 28),            # Sistema Financiero, Seguros, SPP, etc.
        ("Tipo de Empresa", 22),      # Banco, Financiera, Caja, etc.
        ("Direccion", 55),
        ("Representante Legal", 35),
        ("Telefono", 18),
        ("Sitio Web", 35),
        ("Auditor", 50),              # Power Automate llena esto
        ("Fee Auditoria", 18),        # Power Automate llena esto
        ("Tipo de Opinion", 22),      # Power Automate llena esto
        ("Directorio", 50),           # Power Automate llena esto
        ("Fuente", 10),
        ("Fecha Extraccion", 20),
    ]
    for i, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = HF, HFL, HA, BRD
        ws.column_dimensions[get_column_letter(i)].width = w

    ds = sorted(datos, key=lambda x: (x["categoria"], x["tipo_empresa"], x["razon_social"].upper()))
    fecha_extraccion = datetime.now().strftime("%d/%m/%Y %H:%M")

    for i, r in enumerate(ds, 2):
        fill = ALT if i % 2 == 0 else PatternFill()
        values = [
            r["razon_social"],
            r["ruc"],
            r["tipo_sociedad"],
            r["categoria"],
            r["tipo_empresa"],
            r["direccion"],
            r["representante"],
            r["telefono"],
            r["web"],
            r["auditor"],             # Vacio - Power Automate llena
            r["fee"],                 # Vacio - Power Automate llena
            r["tipo_opinion"],        # Vacio - Power Automate llena
            r["directorio_miembros"], # Vacio - Power Automate llena
            "SBS",
            fecha_extraccion,
        ]
        for col, v in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font, c.alignment, c.border, c.fill = CF, CA, BRD, fill

    # Crear tabla para Power Automate
    if len(ds) > 0:
        last_col = get_column_letter(len(headers))  # N (14 columnas)
        table_ref = f"A1:{last_col}{len(ds) + 1}"
        tabla = Table(displayName="tblDirectorioSBS", ref=table_ref)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tabla)
    ws.freeze_panes = "A2"

    # --- Hoja 2: Resumen ---
    ws2 = wb.create_sheet("Resumen")
    total = len(ds)
    tipos = {}
    for r in ds:
        t = r["tipo_empresa"]
        tipos[t] = tipos.get(t, 0) + 1

    metricas = [
        ("Total empresas supervisadas", total),
        ("", ""),
    ]
    for t, c in sorted(tipos.items()):
        metricas.append((t, c))
    metricas.extend([
        ("", ""),
        ("Fuente", "SBS - Superintendencia de Banca, Seguros y AFP"),
        ("Fecha extraccion", fecha_extraccion),
        ("", ""),
        ("Nota", "Columnas Auditor, Fee, Tipo Opinion y Directorio estan vacias."),
        ("", "La SBS no publica PDFs de dictamenes en su portal."),
        ("", "Esos datos se obtienen de los sitios web individuales de cada empresa."),
    ])

    for i, (k, v) in enumerate(metricas, 1):
        ws2.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=i, column=2, value=str(v)).font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 65

    wb.save(filepath)
    log.info(f"\nExcel guardado: {filepath}")
    log.info(f"  Total empresas: {total}")
    for t, c in sorted(tipos.items()):
        log.info(f"  {t}: {c}")


# ============================================================================
# MAIN
# ============================================================================
def main():
    p = argparse.ArgumentParser(description="Scraper SBS v1 - Directorio Empresas Supervisadas")
    p.add_argument("--no-headless", action="store_true", help="Mostrar navegador")
    p.add_argument("--max", type=int, default=None, help="Limitar cantidad de categorias (para pruebas)")
    args = p.parse_args()

    log.info("=" * 60)
    log.info("SCRAPER SBS v1 - Directorio Empresas Supervisadas")
    log.info("=" * 60)

    driver = None
    try:
        driver = crear_driver(headless=not args.no_headless)
        log.info("Chrome OK")
        datos = scrape(driver, args.max)
        if datos:
            excel_path = OUTPUT_DIR / EXCEL_NAME
            exportar(datos, excel_path)
        else:
            log.warning("No se encontraron empresas. Revisa los archivos debug_sbs_*.html en output/")
        try:
            driver.save_screenshot(str(LOG_DIR / f"screenshot_sbs_{TS}.png"))
        except Exception:
            pass
    except Exception as e:
        log.error(f"Error fatal: {e}")
    finally:
        if driver:
            driver.quit()

    log.info("\nScraping SBS finalizado")


if __name__ == "__main__":
    main()
