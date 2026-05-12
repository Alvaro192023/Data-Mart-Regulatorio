"""
============================================================================
SCRAPER BVL vF.2025 - Fix descarga PDFs Memorias 2025
============================================================================
Correcciones vs v5.1:
  - Fix dropdown año: 3 metodos de seleccion (Select nativo, Angular, JS)
    con verificacion de que el año se seleccionó correctamente
  - Estado "Sin dictamen": cuando no se encuentra Memoria Anual para 2025
    (antes quedaba como "OK" sin PDFs)
  - Columnas Excel: quitado "Tipo de Opinion", agregado "Temas de Interes",
    "Sostenibilidad", "Estructura y Solvencia" (se llenan via Power Automate)
  - Toda la extraccion de datos (info corp, indices, etc.) sin cambios

Uso:
    python scraper_bvl_vF_2025.py                  # Completo
    python scraper_bvl_vF_2025.py --max 5          # Prueba
    python scraper_bvl_vF_2025.py --no-headless    # Ver navegador
    python scraper_bvl_vF_2025.py --sin-pdfs       # Sin descarga PDFs
============================================================================
"""
import time, random, re, sys, argparse, logging
from datetime import datetime
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException,
    StaleElementReferenceException, ElementClickInterceptedException
)
import requests as req

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================================
# CONFIGURACION
# ============================================================================
URL_LISTADO = "https://www.bvl.com.pe/emisores/listado-emisores"
URL_DETALLE = "https://www.bvl.com.pe/emisores/detalle?companyCode={code}"
DELAY = (0.5, 1.5)
MAX_REINTENTOS = 3
MIN_PDF_SIZE = 1024
EXCEL_NAME = "bvl_emisores.xlsx"
ANIOS_MEMORIAS = ["2025"]  # Solo 2025

INDICES_FINANCIEROS = [
    "Liquidez", "Liquidez (%)", "Rotación en Activos", "Solvencia",
    "Deuda / Patrimonio", "Rentabilidad de Capital %",
    "Rentabilidad de Patrimonio %", "Valor en libros %"
]
ANIOS_INDICES = ["2024", "2023", "2022", "2021"]

#SHAREPOINT_DIR = Path(r"C:\Users\avillanuev044\OneDrive - PwC\PE-IFS-SALES - Entidades de Regulación\BVL")
SHAREPOINT_DIR = Path("output")
OUTPUT_DIR = SHAREPOINT_DIR
PDF_DIR = SHAREPOINT_DIR / "pdfs_bvl"
LOG_DIR = Path("output")
TS = datetime.now().strftime("%Y%m%d_%H%M%S")

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
PDF_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_DIR / f"scraper_bvl_{TS}.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

SESSION = req.Session()
SESSION.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})

# ============================================================================
# UTILIDADES
# ============================================================================
def limpiar_carpeta(nombre):
    nombre = re.sub(r'[<>:"/\\|?*(){}[\]#%&!]', '', nombre)
    return re.sub(r'\s+', ' ', nombre.strip()).rstrip('. ')[:80].strip()

def tipo_sociedad(nombre):
    for p, t in [(r'\bS\.?\s*A\.?\s*A\.?\b','S.A.A.'),(r'\bS\.?\s*A\.?\s*C\.?\b','S.A.C.'),
        (r'\bS\.?\s*A\.?\b(?!\.\s*[AC])','S.A.'),(r'\bS\.?\s*Civil\b','S. Civil'),
        (r'\bS\.?\s*R\.?\s*L\.?\b','S.R.L.'),(r'\bSUCURSAL\b','Sucursal')]:
        if re.search(p, nombre, re.IGNORECASE): return t
    return ""

def crear_driver(headless=True):
    opts = Options()
    if headless: opts.add_argument("--headless=new")
    for a in ["--no-sandbox","--disable-dev-shm-usage","--disable-gpu","--window-size=1920,1080",
              "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
              "--disable-blink-features=AutomationControlled"]:
        opts.add_argument(a)
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        svc = ChromeService(ChromeDriverManager().install())
    except ImportError:
        svc = ChromeService(r"C:\Users\avillanuev044\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe")
    d = webdriver.Chrome(service=svc, options=opts)
    d.set_page_load_timeout(60); d.implicitly_wait(3)
    return d

def esperar(driver, t=15):
    try:
        WebDriverWait(driver, t).until(lambda d: d.execute_script("return document.readyState") == "complete")
        time.sleep(1.5)
    except: pass

def cookies(driver):
    try:
        for b in driver.find_elements(By.CSS_SELECTOR, "bvl-cookies-banner button.btn-nuam"):
            if b.is_displayed(): b.click(); time.sleep(1); return
        for b in driver.find_elements(By.XPATH, "//button[contains(text(),'Aceptar')]"):
            if b.is_displayed(): b.click(); time.sleep(1); return
    except: pass

def click_tab(driver, nombre):
    for sel in [f"//a[normalize-space(text())='{nombre}']",
                f"//a[contains(normalize-space(text()),'{nombre}')]",
                f"//*[contains(@class,'tab')]//a[contains(text(),'{nombre.split()[0]}')]"]:
        try:
            for el in driver.find_elements(By.XPATH, sel):
                if el.is_displayed():
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    time.sleep(0.3)
                    try: el.click()
                    except: driver.execute_script("arguments[0].click();", el)
                    time.sleep(1.5); return True
        except: continue
    return False

def descargar_pdf(url, ruta):
    for i in range(1, MAX_REINTENTOS+1):
        try:
            r = SESSION.get(url, timeout=45, allow_redirects=True)
            if r.status_code != 200: continue
            if r.content[:5] != b'%PDF-': return False
            if len(r.content) < MIN_PDF_SIZE: return False
            ruta.write_bytes(r.content)
            log.info(f"      PDF: {ruta.name} ({len(r.content)//1024} KB)")
            return True
        except:
            if i < MAX_REINTENTOS: time.sleep(3*i)
    return False

# ============================================================================
# FASE 1: LISTADO (identico a v3)
# ============================================================================
def listado_completo(driver):
    log.info("Navegando a listado de emisores BVL...")
    driver.get(URL_LISTADO); esperar(driver, 30); cookies(driver); time.sleep(2)
    try:
        b = driver.find_element(By.CSS_SELECTOR, "button[ctrl='linkAll']")
        if "active" not in (b.get_attribute("class") or ""): b.click(); time.sleep(3)
    except: pass

    todos, seen = [], set()
    def add(pag):
        n=0
        for e in pag:
            if e["c"] not in seen: seen.add(e["c"]); todos.append(e); n+=1
        return n
    def ext():
        r=[]
        for f in driver.find_elements(By.CSS_SELECTOR, "tr.bvl-row"):
            try:
                lk = f.find_element(By.CSS_SELECTOR, "a[ctrl='columnCompany']")
                hr = lk.get_attribute("href") or ""
                m = re.search(r'companyCode=([^&\s]+)', hr)
                if not m: continue
                c = m.group(1).strip()
                n = (lk.get_attribute("title") or lk.text).strip()
                if not n: continue
                s = ""
                try:
                    sp = f.find_elements(By.CSS_SELECTOR, "td span[title]")
                    if sp: s = sp[0].get_attribute("title").strip()
                except: pass
                r.append({"c": c, "n": n, "s": s})
            except: continue
        return r

    add(ext()); log.info(f"  Pagina 1: {len(todos)} emisores")
    tp = 1
    try:
        v = driver.find_element(By.CSS_SELECTOR, "bvl-pagination").get_attribute("ng-reflect-total-pages")
        if v: tp = int(v)
    except:
        try:
            ns = [int(b.text.strip()) for b in driver.find_elements(By.CSS_SELECTOR, "button[ctrl='paginationItem']") if b.text.strip().isdigit()]
            if ns: tp = max(ns)
        except: pass
    log.info(f"  Total paginas: {tp}")

    for p in range(2, tp+1):
        ok = False
        for b in driver.find_elements(By.CSS_SELECTOR, "button[ctrl='paginationItem']"):
            if b.text.strip() == str(p):
                try: driver.execute_script("arguments[0].scrollIntoView({block:'center'});", b); time.sleep(0.3); b.click()
                except: driver.execute_script("arguments[0].click();", b)
                ok=True; break
        if not ok:
            try: nx = driver.find_element(By.CSS_SELECTOR, "button[ctrl='nextPagination']"); driver.execute_script("arguments[0].click();", nx); ok=True
            except: break
        time.sleep(3)
        try:
            WebDriverWait(driver,10).until(lambda d: any(
                b.text.strip()==str(p) and "active" in (b.get_attribute("class") or "")
                for b in d.find_elements(By.CSS_SELECTOR, "button[ctrl='paginationItem']")))
        except: pass
        n = add(ext()); log.info(f"  Pagina {p}/{tp}: +{n} (total: {len(todos)})")
    return todos

# ============================================================================
# FASE 2: EXTRAER DETALLE CON JS INJECTION
# ============================================================================
JS_INFO_CORP = """
var data = {razon_social:'', telefono:'', direccion:'', pagina_web:'', fundacion:'',
            descripcion:'', directorio:'', gerente_general:''};

// 1. Info Corporativa - buscar en DIV.col-md-9 el texto estructurado
var infoDiv = document.querySelector('.col-md-9');
if (infoDiv) {
    var txt = infoDiv.innerText;
    var lines = txt.split('\\n').map(l => l.trim()).filter(l => l.length > 0);
    for (var i = 0; i < lines.length; i++) {
        var line = lines[i];
        if (line.match(/^Raz[oó]n social$/i) && i+1 < lines.length) data.razon_social = lines[i+1];
        else if (line.match(/^Tel[eé]fono/i) && i+1 < lines.length) data.telefono = lines[i+1];
        else if (line.match(/^Direcci[oó]n$/i) && i+1 < lines.length) data.direccion = lines[i+1];
        else if (line.match(/^P[aá]gina Web$/i) && i+1 < lines.length) data.pagina_web = lines[i+1];
        else if (line.match(/^Fundaci[oó]n$/i) && i+1 < lines.length) data.fundacion = lines[i+1];
    }
}

// 2. RUC - no se extrae aqui, se completa con padron SUNAT

// 3. Descripcion - buscar el parrafo bajo H2 "Descripcion"
var h2s = document.querySelectorAll('h2');
for (var h of h2s) {
    if (h.innerText.toLowerCase().includes('descripci')) {
        var next = h.nextElementSibling;
        while (next) {
            var t = next.innerText.trim();
            if (t.length > 30) { data.descripcion = t.substring(0, 500); break; }
            next = next.nextElementSibling;
        }
        break;
    }
}

// 4. Directorio - extraer de SECTION.directory
var dirSections = document.querySelectorAll('section.directory');
var dirParts = [];
var gerenteFound = false;
for (var sec of dirSections) {
    var strongs = sec.querySelectorAll('strong');
    for (var s of strongs) {
        var parentText = s.parentElement ? s.parentElement.innerText.trim() : s.innerText.trim();
        if (parentText.match(/^Gerente General/i)) {
            // Es la seccion de gerente
            var lines2 = parentText.split('\\n').map(l=>l.trim()).filter(l=>l.length>0);
            for (var l of lines2) {
                if (l.match(/^Gerente General:/i)) {
                    data.gerente_general = l.replace(/^Gerente General:\\s*/i, '').trim();
                    break;
                }
            }
        } else if (parentText.match(/^(Presidente|Director|Vice)/i)) {
            dirParts.push(parentText);
        }
    }
}
if (dirParts.length === 0) {
    // Fallback: buscar en todo el body
    var bodyText = document.body.innerText;
    var dirMatch = bodyText.match(/Directorio[\\s\\S]*?(?=Gerente General|Lleve su empresa|$)/);
    if (dirMatch) {
        var dirLines = dirMatch[0].split('\\n').map(l=>l.trim()).filter(l=>l.length>3);
        for (var dl of dirLines) {
            if (dl.match(/^(Presidente|Director|Vice)/i)) dirParts.push(dl);
        }
    }
    // Gerente fallback
    if (!data.gerente_general) {
        var gerMatch = bodyText.match(/Gerente General:\\s*([^\\n]+)/i);
        if (gerMatch) data.gerente_general = gerMatch[1].trim();
    }
}
data.directorio = dirParts.join(' | ');

return JSON.stringify(data);
"""

def extraer_info_corporativa_js(driver):
    """Extrae info corporativa usando JavaScript injection."""
    try:
        result = driver.execute_script(JS_INFO_CORP)
        import json
        return json.loads(result)
    except Exception as e:
        log.debug(f"      Error JS info corp: {e}")
        return {}

def extraer_valores_inscritos(driver):
    vals = ""
    if not click_tab(driver, "Valores Inscritos"): return vals
    time.sleep(1)
    try:
        body = driver.find_element(By.TAG_NAME, "body").text
        nemos = re.findall(r'\b([A-Z]{3,}[A-Z0-9]*[CI]1)\b', body)
        if nemos: vals = ", ".join(sorted(set(nemos)))
    except: pass
    return vals

def extraer_indices_financieros(driver):
    indices = {}
    if not click_tab(driver, "Información financiera"):
        if not click_tab(driver, "financiera"): return indices
    time.sleep(1.5)
    try:
        for tabla in driver.find_elements(By.TAG_NAME, "table"):
            headers = tabla.find_elements(By.TAG_NAME, "th")
            anios = [h.text.strip() for h in headers if re.match(r'^\d{4}$', h.text.strip())]
            if not anios: continue
            for fila in tabla.find_elements(By.TAG_NAME, "tr"):
                celdas = fila.find_elements(By.TAG_NAME, "td")
                if len(celdas) < 2: continue
                nombre = celdas[0].text.strip()
                if not nombre: continue
                for i, anio in enumerate(anios):
                    ci = i + 1
                    if ci < len(celdas):
                        val = celdas[ci].text.strip()
                        if val and val != "--": indices[f"{nombre}_{anio}"] = val
            if indices: break
        if not indices:
            for tabla in driver.find_elements(By.TAG_NAME, "table"):
                filas = tabla.find_elements(By.TAG_NAME, "tr")
                if len(filas) < 2: continue
                hc = filas[0].find_elements(By.CSS_SELECTOR, "td, th")
                anios = [h.text.strip() for h in hc if re.match(r'^\d{4}$', h.text.strip())]
                if not anios: continue
                for fila in filas[1:]:
                    celdas = fila.find_elements(By.TAG_NAME, "td")
                    if len(celdas) < 2: continue
                    nombre = celdas[0].text.strip()
                    if not nombre: continue
                    for i, anio in enumerate(anios):
                        ci = i+1
                        if ci < len(celdas):
                            val = celdas[ci].text.strip()
                            if val and val != "--": indices[f"{nombre}_{anio}"] = val
                if indices: break
    except Exception as e:
        log.debug(f"      Error indices: {e}")
    return indices

def seleccionar_anio_memorias(driver, anio):
    """Selecciona un año en el dropdown de Memorias. Retorna True si tuvo exito."""

    # Desactivar implicit wait para no esperar 3s por cada selector que no existe
    driver.implicitly_wait(0)

    try:
        # Metodo 0: verificar si el año ya está seleccionado
        try:
            page_text = driver.find_element(By.TAG_NAME, "body").text
            if re.search(rf'\d{{2}}/\d{{2}}/{str(int(anio)+1)}', page_text):
                log.info(f"      Año {anio} ya visible en la pagina")
                return True
        except: pass

        # Metodo 1: <select> nativo
        for sel in driver.find_elements(By.TAG_NAME, "select"):
            try:
                if not sel.is_displayed(): continue
                opciones = sel.find_elements(By.TAG_NAME, "option")
                opciones_text = [o.text.strip() for o in opciones]
                if anio in opciones_text:
                    from selenium.webdriver.support.ui import Select
                    Select(sel).select_by_visible_text(anio)
                    time.sleep(1.5)
                    log.info(f"      Año {anio} seleccionado (select nativo)")
                    return True
            except: continue

        # Metodo 2: Angular dropdown
        for sel_css in ["bvl-select", ".g-site-select", "[class*='select']"]:
            for dd in driver.find_elements(By.CSS_SELECTOR, sel_css):
                try:
                    if not dd.is_displayed(): continue
                    dd_text = dd.text.strip()
                    if not any(a in dd_text for a in ["2024","2025","2026"]): continue
                    if anio in dd_text:
                        log.info(f"      Año {anio} ya seleccionado")
                        return True
                    driver.execute_script("arguments[0].click();", dd)
                    time.sleep(0.5)
                    for opt in driver.find_elements(By.XPATH, f"//*[normalize-space(text())='{anio}']"):
                        if opt.is_displayed():
                            driver.execute_script("arguments[0].click();", opt)
                            time.sleep(1.5)
                            log.info(f"      Año {anio} seleccionado (Angular)")
                            return True
                except: continue

        # Metodo 3: JavaScript
        try:
            result = driver.execute_script(f"""
                var selects = document.querySelectorAll('select');
                for (var s of selects) {{
                    for (var o of s.options) {{
                        if (o.text.trim() === '{anio}' || o.value === '{anio}') {{
                            s.value = o.value;
                            s.dispatchEvent(new Event('change', {{bubbles: true}}));
                            return 'select';
                        }}
                    }}
                }}
                return false;
            """)
            if result:
                time.sleep(1.5)
                log.info(f"      Año {anio} OK (JS)")
                return True
        except: pass

        log.info(f"      Dropdown año: usando vista por defecto")
        return False

    finally:
        # Restaurar implicit wait
        driver.implicitly_wait(3)


def extraer_pdfs_memorias(driver, carpeta):
    """Descarga PDFs de Memorias 2025. Retorna (lista_pdfs, tiene_memoria_anual).
    Nombre de archivo: {titulo}_{DDMMYYYY}.pdf (fecha de la tabla BVL).
    FILTRO: Las Memorias del año N se publican en N+1 (ej: Memorias 2025 → fecha 2026).
    Solo descarga PDFs cuya fecha de publicacion sea del año correcto."""
    pdfs = []
    tiene_memoria_anual = False
    urls_descargadas = set()

    if not click_tab(driver, "Memorias"):
        log.warning("      No se pudo hacer click en tab Memorias")
        return pdfs, tiene_memoria_anual
    time.sleep(1)

    for anio in ANIOS_MEMORIAS:
        # Las memorias del año N se publican en N+1
        anio_publicacion = str(int(anio) + 1)
        log.info(f"      Buscando Memorias {anio} (publicadas en {anio_publicacion})...")
        seleccionar_anio_memorias(driver, anio)
        time.sleep(1)

        # Recolectar links con fecha desde filas de tabla
        links_encontrados = []  # (href, titulo, fecha_str)
        seen_hrefs = set()

        for fila in driver.find_elements(By.CSS_SELECTOR, "tr"):
            try:
                links_en_fila = fila.find_elements(By.TAG_NAME, "a")
                for link in links_en_fila:
                    href = link.get_attribute("href") or ""
                    if "documents.bvl.com.pe" not in href: continue
                    if not href.lower().endswith(".pdf"): continue
                    if href in seen_hrefs: continue
                    seen_hrefs.add(href)
                    titulo = link.text.strip()
                    if not titulo or len(titulo) < 3: continue
                    if titulo.lower() in ["leer más","ver más","aceptar"]: continue

                    # Extraer fecha de la fila (DD/MM/YYYY)
                    fecha_str = ""
                    fecha_anio = ""
                    texto_fila = fila.text
                    fecha_match = re.search(r'(\d{2})/(\d{2})/(\d{4})', texto_fila)
                    if fecha_match:
                        dd, mm, yyyy = fecha_match.group(1), fecha_match.group(2), fecha_match.group(3)
                        fecha_str = f"{dd}{mm}{yyyy}"
                        fecha_anio = yyyy

                    # FILTRO: solo PDFs publicados en el año correcto (N+1)
                    if fecha_anio and fecha_anio != anio_publicacion:
                        continue  # PDF de otro año, saltar

                    links_encontrados.append((href, titulo, fecha_str))
            except StaleElementReferenceException: continue
            except: continue

        log.info(f"      {len(links_encontrados)} PDFs del {anio_publicacion} encontrados (de {len(seen_hrefs)} links totales)")

        # Descargar
        for href, titulo, fecha in links_encontrados:
            if href in urls_descargadas:
                continue
            urls_descargadas.add(href)

            titulo_limpio = re.sub(r'[<>:"/\\|?*]', '', titulo)
            titulo_limpio = re.sub(r'\s+', '_', titulo_limpio.strip())[:70]

            if fecha:
                pdf_nombre = f"{titulo_limpio}_{fecha}.pdf"
            else:
                pdf_nombre = f"{titulo_limpio}.pdf"

            ruta = carpeta / pdf_nombre

            if ruta.exists():
                log.info(f"      Ya existe: {pdf_nombre}, saltando")
                pdfs.append(pdf_nombre)
                if "Memoria_Anual" in pdf_nombre:
                    tiene_memoria_anual = True
                continue

            if descargar_pdf(href, ruta):
                pdfs.append(pdf_nombre)
                if "Memoria_Anual" in pdf_nombre:
                    tiene_memoria_anual = True

    return pdfs, tiene_memoria_anual

    return pdfs, tiene_memoria_anual


# ============================================================================
# SCRAPING POR EMPRESA
# ============================================================================
def detalle_empresa(driver, em, idx, total, con_pdfs=True):
    codigo, nombre, sector = em["c"], em["n"], em["s"]
    log.info(f"[{idx}/{total}] {nombre[:55]} ({codigo})")

    r = {
        "razon_social": nombre, "codigo_bvl": codigo, "sector": sector,
        "tipo_sociedad": tipo_sociedad(nombre),
        "ruc":"", "telefono":"", "direccion":"", "pagina_web":"", "fundacion":"",
        "descripcion":"", "directorio":"", "gerente_general":"",
        "nemonicos":"", "carpeta": limpiar_carpeta(nombre),
        "auditor":"","fee":"",
        "temas_interes":"","sostenibilidad":"","estructura_solvencia":"",
        "fuente":"BVL","estado":"OK",
        "pdfs_descargados":0, "pdfs_nombres":"",
    }
    for idx_n in INDICES_FINANCIEROS:
        for a in ANIOS_INDICES:
            r[f"{idx_n}_{a}"] = ""

    for intento in range(1, MAX_REINTENTOS+1):
        try:
            driver.get(URL_DETALLE.format(code=codigo)); esperar(driver); cookies(driver)

            # 1. Info Corporativa + Directorio + Gerente (JS injection)
            log.info(f"    Info corporativa + Directorio...")
            info = extraer_info_corporativa_js(driver)
            if info:
                for k in ["telefono","direccion","pagina_web","fundacion","descripcion","directorio","gerente_general"]:
                    if info.get(k): r[k] = info[k]

            # 2. Valores Inscritos
            log.info(f"    Valores inscritos...")
            r["nemonicos"] = extraer_valores_inscritos(driver)

            # 3. Indices Financieros
            log.info(f"    Indices financieros...")
            indices = extraer_indices_financieros(driver)
            if indices:
                r.update(indices)
                log.info(f"      {len(indices)} valores")
            else:
                log.info(f"      Sin datos")

            # 4. PDFs Memorias 2025
            if con_pdfs:
                log.info(f"    Memorias PDFs 2025...")
                carp = PDF_DIR / r["carpeta"]; carp.mkdir(parents=True, exist_ok=True)
                pdfs, tiene_memoria_anual = extraer_pdfs_memorias(driver, carp)
                r["pdfs_descargados"] = len(pdfs)
                r["pdfs_nombres"] = " | ".join(pdfs)
                if pdfs:
                    log.info(f"      {len(pdfs)} PDFs")
                    if tiene_memoria_anual:
                        log.info(f"      ✓ Memoria Anual encontrada")
                    else:
                        log.info(f"      ⚠ PDFs descargados pero sin Memoria Anual")
                        r["estado"] = "Sin dictamen"
                else:
                    log.info(f"      Sin PDFs → Sin dictamen")
                    r["estado"] = "Sin dictamen"
                    if carp.exists() and not any(carp.iterdir()):
                        try: carp.rmdir()
                        except: pass
            break
        except TimeoutException:
            if intento < MAX_REINTENTOS: time.sleep(5*intento)
            else: r["estado"] = "Error"
        except Exception as e:
            em_str = str(e)
            if "net::" in em_str or "ERR_" in em_str:
                if intento < MAX_REINTENTOS: time.sleep(10)
                else: r["estado"] = "Error"
            else: log.error(f"    Error: {em_str[:80]}"); r["estado"] = "Error"; break

    time.sleep(random.uniform(*DELAY))
    return r


# ============================================================================
# SCRAPING PRINCIPAL
# ============================================================================
def scrape(driver, max_emp=None, solo_listado=False, sin_pdfs=False):
    emisores = listado_completo(driver)
    if not emisores: return []
    log.info(f"Total emisores: {len(emisores)}")
    if max_emp: emisores = emisores[:max_emp]; log.info(f"Limitado a {max_emp}")
    if solo_listado:
        f = datetime.now().strftime("%d/%m/%Y %H:%M")
        return [{"razon_social":e["n"],"codigo_bvl":e["c"],"sector":e["s"],
                 "tipo_sociedad":tipo_sociedad(e["n"]),"carpeta":limpiar_carpeta(e["n"]),
                 "fuente":"BVL","estado":"OK","fecha_extraccion":f} for e in emisores]

    log.info(f"\nExtrayendo detalle de {len(emisores)} emisores...")
    resultados = []
    for idx, em in enumerate(emisores, 1):
        r = detalle_empresa(driver, em, idx, len(emisores), con_pdfs=not sin_pdfs)
        r["fecha_extraccion"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        resultados.append(r)
    return resultados


# ============================================================================
# EXCEL
# ============================================================================
def exportar(datos, filepath):
    wb = Workbook()
    HF = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HFL = PatternFill("solid", fgColor="1F4E79")
    HA = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CF = Font(name="Arial", size=10)
    CA = Alignment(vertical="top", wrap_text=True)
    BRD = Border(left=Side("thin","B0B0B0"),right=Side("thin","B0B0B0"),
                 top=Side("thin","B0B0B0"),bottom=Side("thin","B0B0B0"))
    ALT = PatternFill("solid", fgColor="E8F0FE")
    ERR = PatternFill("solid", fgColor="FCE4EC")
    FIN = PatternFill("solid", fgColor="E8F4E8")
    SIN_DICT = PatternFill("solid", fgColor="FFF3CD")  # Amarillo para Sin dictamen

    ws = wb.active; ws.title = "Emisores_BVL"

    headers = [
        ("Razon Social",50),("RUC",16),("Tipo Sociedad",13),("Codigo BVL",14),
        ("Sector",25),("Telefono",16),("Direccion",45),("Pagina Web",30),
        ("Fundacion",14),("Descripcion",60),
        ("Directorio",70),("Gerente General",35),
        ("Nemonicos",25),("Carpeta",35),
        ("Auditor",45),("Fee Auditoria",16),
        ("Temas de Interes",60),("Sostenibilidad",60),("Estructura y Solvencia",60),
        ("Fuente",8),("Estado",12),
        ("PDFs Descargados",8),("Nombres PDFs",55),
        ("Fecha Extraccion",18),
    ]
    col_base = len(headers)

    for idx_n in INDICES_FINANCIEROS:
        for a in ANIOS_INDICES:
            headers.append((f"{idx_n} {a}", 14))

    for i,(h,w) in enumerate(headers,1):
        c = ws.cell(row=1,column=i,value=h)
        c.font,c.fill,c.alignment,c.border = HF,HFL,HA,BRD
        ws.column_dimensions[get_column_letter(i)].width = w

    ds = sorted(datos, key=lambda x: x.get("razon_social","").upper())

    for i,r in enumerate(ds,2):
        es_err = r.get("estado")=="Error"
        es_sin_dict = r.get("estado")=="Sin dictamen"
        fill = ERR if es_err else (SIN_DICT if es_sin_dict else (ALT if i%2==0 else PatternFill()))
        vals = [
            r.get("razon_social",""),r.get("ruc",""),r.get("tipo_sociedad",""),
            r.get("codigo_bvl",""),r.get("sector",""),r.get("telefono",""),
            r.get("direccion",""),r.get("pagina_web",""),r.get("fundacion",""),
            r.get("descripcion",""),r.get("directorio",""),r.get("gerente_general",""),
            r.get("nemonicos",""),r.get("carpeta",""),
            r.get("auditor",""),r.get("fee",""),
            r.get("temas_interes",""),r.get("sostenibilidad",""),r.get("estructura_solvencia",""),
            r.get("fuente","BVL"),r.get("estado",""),
            r.get("pdfs_descargados",0),r.get("pdfs_nombres",""),
            r.get("fecha_extraccion",""),
        ]
        for idx_n in INDICES_FINANCIEROS:
            for a in ANIOS_INDICES:
                vals.append(r.get(f"{idx_n}_{a}",""))

        for col,v in enumerate(vals,1):
            c = ws.cell(row=i,column=col,value=v)
            c.font,c.alignment,c.border = CF,CA,BRD
            if col > col_base:
                c.fill = FIN if not es_err else ERR
                if isinstance(v,str) and v:
                    try: c.value = float(v); c.number_format = '0.000000'
                    except: pass
            else:
                c.fill = fill

    lr = len(ds)+1; lc = get_column_letter(len(headers))
    t = Table(displayName="tblEmisores", ref=f"A1:{lc}{lr}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",showFirstColumn=False,
        showLastColumn=False,showRowStripes=True,showColumnStripes=False)
    ws.add_table(t); ws.freeze_panes = "A2"

    # Resumen
    ws2 = wb.create_sheet("Resumen")
    total=len(ds)
    idx_c=sum(1 for d in ds if any(d.get(f"Liquidez_{a}") for a in ANIOS_INDICES))
    dir_c=sum(1 for d in ds if d.get("directorio"))
    ger_c=sum(1 for d in ds if d.get("gerente_general"))
    pdf_c=sum(1 for d in ds if d.get("pdfs_descargados",0)>0)
    tot_pdf=sum(d.get("pdfs_descargados",0) for d in ds)
    err=sum(1 for d in ds if d.get("estado")=="Error")
    sin_dict=sum(1 for d in ds if d.get("estado")=="Sin dictamen")

    mets=[("Total emisores BVL",total),("Con indices financieros",idx_c),
          ("Con directorio",dir_c),("Con gerente general",ger_c),
          ("Con PDFs Memorias",pdf_c),("Total PDFs",tot_pdf),
          ("Sin dictamen (sin Memoria Anual)",sin_dict),("Errores",err),
          ("",""),("Carpeta PDFs",str(PDF_DIR.resolve())),
          ("",""),("Paso siguiente","Ejecutar flujo Power Automate")]
    for i,(k,v) in enumerate(mets,1):
        ws2.cell(row=i,column=1,value=k).font=Font(name="Arial",bold=True,size=10)
        ws2.cell(row=i,column=2,value=str(v)).font=Font(name="Arial",size=10)
    ws2.column_dimensions["A"].width=35; ws2.column_dimensions["B"].width=60

    wb.save(filepath)
    log.info(f"\nExcel: {filepath}")
    log.info(f"  Total:{total} | Indices:{idx_c} | Dir:{dir_c} | Ger:{ger_c}")
    log.info(f"  PDFs:{tot_pdf} | Sin dictamen:{sin_dict} | Errores:{err}")

# ============================================================================
# MAIN
# ============================================================================
def main():
    p = argparse.ArgumentParser(description="Scraper BVL v5.2")
    p.add_argument("--no-headless", action="store_true")
    p.add_argument("--max", type=int, default=None)
    p.add_argument("--solo-listado", action="store_true")
    p.add_argument("--sin-pdfs", action="store_true")
    args = p.parse_args()

    log.info("="*60)
    log.info("SCRAPER BVL v5.2 - Datos completos + PDFs Memorias 2025")
    log.info("="*60)

    driver = None
    try:
        driver = crear_driver(headless=not args.no_headless)
        log.info("Chrome OK")
        datos = scrape(driver, args.max, args.solo_listado, args.sin_pdfs)
        if datos:
            excel_path = OUTPUT_DIR / EXCEL_NAME
            exportar(datos, excel_path)
        try: driver.save_screenshot(str(LOG_DIR / f"screenshot_bvl_{TS}.png"))
        except: pass
    except Exception as e:
        log.error(f"Error fatal: {e}")
        import traceback; traceback.print_exc()
    finally:
        if driver: driver.quit()
        SESSION.close()

if __name__ == "__main__":
    main()
