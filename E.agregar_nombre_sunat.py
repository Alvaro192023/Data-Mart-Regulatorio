"""
============================================================================
AGREGAR NOMBRE SUNAT - Cruce Consolidado con Padron Reducido SUNAT
============================================================================
Lee el archivo Consolidado_E__Regulación.xlsx y agrega columna
"Razon Social SUNAT" con el nombre oficial de cada empresa en SUNAT,
empresa por empresa.

Estrategia de matching (4 niveles):
  1. Match exacto normalizado (sin acentos, sin puntuación)
  2. Match sin tipo de sociedad (S.A.A., S.A., etc.)
  3. Match sin contenido entre paréntesis ni sufijos
  4. Match fuzzy por primeras N palabras significativas

Requisitos:
    pip install pandas openpyxl requests unidecode

Uso:
    python agregar_nombre_sunat.py
    python agregar_nombre_sunat.py --padron padron_reducido_ruc.txt
    python agregar_nombre_sunat.py --excel otro_archivo.xlsx
============================================================================
"""
import os, sys, re, zipfile, argparse, logging
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

PADRON_ZIP_URL = "http://www2.sunat.gob.pe/padron_reducido_ruc.zip"
CACHE_DIR = Path("output")
CACHE_DIR.mkdir(exist_ok=True)


def descargar_padron():
    import requests
    log.info("Descargando Padron Reducido SUNAT (~200 MB)...")
    log.info(f"  URL: {PADRON_ZIP_URL}")
    zip_path = CACHE_DIR / "padron_reducido_ruc.zip"
    r = requests.get(PADRON_ZIP_URL, stream=True, timeout=600)
    total = int(r.headers.get("content-length", 0))
    dl = 0
    with open(zip_path, "wb") as f:
        for chunk in r.iter_content(1024*1024):
            f.write(chunk); dl += len(chunk)
            if total: print(f"\r  {dl//1024//1024} MB / {total//1024//1024} MB ({dl*100//total}%)", end="")
    print()
    log.info("  Extrayendo...")
    with zipfile.ZipFile(zip_path, "r") as z:
        z.extractall(CACHE_DIR)
        for n in z.namelist():
            if n.endswith(".txt"): return CACHE_DIR / n
    for f in CACHE_DIR.glob("padron*.txt"): return f
    return None


def normalizar(nombre):
    """Normaliza un nombre para matching."""
    if not isinstance(nombre, str): return ""
    try:
        from unidecode import unidecode
        n = unidecode(nombre)
    except ImportError:
        n = nombre
        for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),
                      ("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("ñ","n"),("Ñ","N")]:
            n = n.replace(a, b)
    n = n.upper().strip()
    n = re.sub(r'[^A-Z0-9\s]', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def limpiar_nombre_bvl(nombre):
    """Limpia nombre de BVL/SMV/SBS para matching."""
    if not nombre: return ""
    n = nombre.strip()
    # Remover contenido entre parentesis
    n = re.sub(r'\(.*?\)', '', n)
    # Remover sufijos despues de guion
    n = re.sub(r'\s*-\s*(EN\s+)?(LIQUIDACI[OÓ]N|REESTRUCTURACI[OÓ]N|COFIDE|BANBIF|INTERBANK).*', '', n, flags=re.IGNORECASE)
    return normalizar(n)


def quitar_tipo_sociedad(nombre_norm):
    """Remueve S.A.A., S.A., S.A.C., etc. de un nombre normalizado."""
    n = re.sub(r'\b(SAA|SA|SAC|SRL|SCRL|S A A|S A C|S A|S R L)\b', '', nombre_norm)
    return re.sub(r'\s+', ' ', n).strip()


def cargar_padron(txt_path):
    """Carga el padron en un diccionario optimizado."""
    import pandas as pd
    log.info(f"Cargando padron: {txt_path.name} ({txt_path.stat().st_size//1024//1024} MB)")

    try:
        df = pd.read_csv(txt_path, sep="|", header=None, usecols=[0, 1, 2, 3],
                         names=["ruc","razon_social","estado","condicion"],
                         dtype=str, encoding="latin-1", on_bad_lines="skip")
    except TypeError:
        df = pd.read_csv(txt_path, sep="|", header=None, usecols=[0, 1, 2, 3],
                         names=["ruc","razon_social","estado","condicion"],
                         dtype=str, encoding="latin-1", error_bad_lines=False)

    # Solo personas juridicas
    df = df[df["ruc"].str.startswith("20", na=False)].copy()
    log.info(f"  Personas juridicas: {len(df):,}")

    # Priorizar activos
    df["prioridad"] = df["estado"].apply(lambda x: 0 if str(x).upper()=="ACTIVO" else 1)
    df = df.sort_values("prioridad")

    # Crear indices de busqueda
    padron = {}  # nombre_normalizado -> (ruc, razon_social_original)
    padron_sin_tipo = {}  # sin tipo sociedad -> (ruc, razon_social_original)
    padron_prefijo = {}  # primeras 3 palabras -> lista de (ruc, razon_social, nombre_completo_norm)

    log.info("  Construyendo indices de busqueda...")
    for _, row in df.iterrows():
        rs = str(row["razon_social"]).strip() if row["razon_social"] else ""
        ruc = str(row["ruc"]).strip()
        if not rs: continue

        norm = normalizar(rs)
        if norm and norm not in padron:
            padron[norm] = (ruc, rs)

        sin_tipo = quitar_tipo_sociedad(norm)
        if sin_tipo and sin_tipo not in padron_sin_tipo:
            padron_sin_tipo[sin_tipo] = (ruc, rs)

        palabras = norm.split()
        if len(palabras) >= 2:
            pref = " ".join(palabras[:3]) if len(palabras) >= 3 else " ".join(palabras[:2])
            if pref not in padron_prefijo:
                padron_prefijo[pref] = []
            if len(padron_prefijo[pref]) < 10:  # Limitar para performance
                padron_prefijo[pref].append((ruc, rs, norm))

    log.info(f"  Indice exacto: {len(padron):,}")
    log.info(f"  Indice sin tipo: {len(padron_sin_tipo):,}")
    log.info(f"  Indice prefijo: {len(padron_prefijo):,}")

    return padron, padron_sin_tipo, padron_prefijo


def buscar_en_sunat(nombre_original, padron, padron_sin_tipo, padron_prefijo):
    """Busca una empresa en los indices SUNAT. Retorna (ruc, razon_social_sunat, metodo)."""

    norm = limpiar_nombre_bvl(nombre_original)
    if not norm: return None, None, None

    # 1. Match exacto
    if norm in padron:
        ruc, rs = padron[norm]
        return ruc, rs, "exacto"

    # 2. Match exacto con nombre original (sin limpiar parentesis)
    norm_orig = normalizar(nombre_original)
    if norm_orig in padron:
        ruc, rs = padron[norm_orig]
        return ruc, rs, "exacto_orig"

    # 3. Sin tipo de sociedad
    sin_tipo = quitar_tipo_sociedad(norm)
    if sin_tipo and sin_tipo in padron_sin_tipo:
        ruc, rs = padron_sin_tipo[sin_tipo]
        return ruc, rs, "sin_tipo"

    # 4. Match por prefijo (primeras 3 palabras)
    palabras = norm.split()
    for n_palabras in [3, 2]:
        if len(palabras) >= n_palabras:
            pref = " ".join(palabras[:n_palabras])
            if pref in padron_prefijo:
                candidatos = padron_prefijo[pref]
                # Buscar el mejor match
                mejor = None
                mejor_score = 0
                for ruc, rs, cn in candidatos:
                    # Score: cuantas palabras coinciden
                    palabras_c = set(cn.split())
                    palabras_b = set(norm.split())
                    interseccion = len(palabras_c & palabras_b)
                    score = interseccion / max(len(palabras_b), 1)
                    if score > mejor_score:
                        mejor_score = score
                        mejor = (ruc, rs)
                if mejor and mejor_score >= 0.5:
                    return mejor[0], mejor[1], f"prefijo_{n_palabras}p"

    # 5. Buscar nombre contenido (para nombres cortos como "INTERBANK")
    if len(norm) >= 6:
        for key, (ruc, rs) in padron.items():
            if norm in key:
                return ruc, rs, "contenido"

    return None, None, None


def procesar_excel(excel_path, padron, padron_sin_tipo, padron_prefijo):
    """Lee el Excel, busca cada empresa en SUNAT y agrega columna."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    log.info(f"\nProcesando: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    # Encontrar columna de Razon Social
    col_razon = None
    for col in range(1, ws.max_column+1):
        v = ws.cell(row=1, column=col).value
        if v and "razon" in str(v).lower():
            col_razon = col; break
    if not col_razon:
        col_razon = 1  # Asumir primera columna

    # Agregar columna "Razon Social SUNAT" despues de la ultima
    col_nueva = ws.max_column + 1
    col_ruc = col_nueva + 1
    col_metodo = col_nueva + 2

    # Headers
    HF = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HFL = PatternFill("solid", fgColor="1F4E79")
    HA = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BRD = Border(left=Side("thin","B0B0B0"),right=Side("thin","B0B0B0"),
                 top=Side("thin","B0B0B0"),bottom=Side("thin","B0B0B0"))

    for col, (titulo, ancho) in enumerate(
        [("Razon Social SUNAT", 55), ("RUC SUNAT", 16), ("Metodo Match", 14)], col_nueva):
        c = ws.cell(row=1, column=col, value=titulo)
        c.font = HF; c.fill = HFL; c.alignment = HA; c.border = BRD
        ws.column_dimensions[chr(64+col) if col <= 26 else "A" + chr(64+col-26)].width = ancho

    # Procesar cada empresa
    encontrados = 0
    no_encontrados = 0
    total = 0

    OK_FILL = PatternFill("solid", fgColor="E8F4E8")
    NO_FILL = PatternFill("solid", fgColor="FCE4EC")
    CF = Font(name="Arial", size=10)

    for row in range(2, ws.max_row+1):
        nombre = ws.cell(row=row, column=col_razon).value
        if not nombre: continue
        nombre = str(nombre).strip()
        total += 1

        ruc, rs_sunat, metodo = buscar_en_sunat(nombre, padron, padron_sin_tipo, padron_prefijo)

        if rs_sunat:
            encontrados += 1
            fill = OK_FILL
            log.info(f"  [{total}] OK ({metodo}): {nombre[:45]} -> {rs_sunat[:45]}")
        else:
            no_encontrados += 1
            rs_sunat = ""
            ruc = ""
            metodo = "NO ENCONTRADO"
            fill = NO_FILL
            log.info(f"  [{total}] ?? : {nombre[:60]}")

        ws.cell(row=row, column=col_nueva, value=rs_sunat).font = CF
        ws.cell(row=row, column=col_nueva).fill = fill
        ws.cell(row=row, column=col_nueva).border = BRD
        ws.cell(row=row, column=col_ruc, value=ruc).font = CF
        ws.cell(row=row, column=col_ruc).fill = fill
        ws.cell(row=row, column=col_ruc).border = BRD
        ws.cell(row=row, column=col_metodo, value=metodo).font = CF
        ws.cell(row=row, column=col_metodo).fill = fill
        ws.cell(row=row, column=col_metodo).border = BRD

    # Guardar
    wb.save(excel_path)
    log.info(f"\n{'='*60}")
    log.info(f"RESULTADO")
    log.info(f"{'='*60}")
    log.info(f"  Total empresas:  {total}")
    log.info(f"  Encontradas:     {encontrados} ({encontrados*100//total}%)")
    log.info(f"  No encontradas:  {no_encontrados}")
    log.info(f"  Excel guardado:  {excel_path}")


def main():
    p = argparse.ArgumentParser(description="Agregar nombre SUNAT al consolidado")
    p.add_argument("--padron", type=str, default=None, help="Ruta al TXT del padron")
    p.add_argument("--excel", type=str, default=None, help="Ruta al Excel consolidado")
    args = p.parse_args()

    log.info("="*60)
    log.info("AGREGAR NOMBRE SUNAT - Cruce con Padron Reducido")
    log.info("="*60)

    # Excel
    if args.excel:
        excel_path = Path(args.excel)
    else:
        # Buscar en ubicaciones comunes
        for p in [Path("Consolidado_E__Regulación.xlsx"),
                  Path("output/Consolidado_E__Regulación.xlsx"),
                  Path(r"C:\Users\Alvaro\Downloads\Consolidado E. Regulación.xlsx")]:
            if p.exists(): excel_path = p; break
        else:
            log.error("No se encontro el Excel. Usa --excel ruta_al_archivo.xlsx")
            sys.exit(1)

    # Padron
    padron_txt = None
    if args.padron:
        padron_txt = Path(args.padron)
    else:
        for f in CACHE_DIR.glob("padron*.txt"):
            padron_txt = f; break
        if not padron_txt:
            padron_txt = descargar_padron()
            if not padron_txt:
                log.error("No se pudo obtener el padron"); sys.exit(1)

    if not padron_txt.exists():
        log.error(f"Padron no encontrado: {padron_txt}"); sys.exit(1)

    # Cargar padron
    padron, padron_sin_tipo, padron_prefijo = cargar_padron(padron_txt)

    # Procesar
    procesar_excel(excel_path, padron, padron_sin_tipo, padron_prefijo)


if __name__ == "__main__":
    main()
