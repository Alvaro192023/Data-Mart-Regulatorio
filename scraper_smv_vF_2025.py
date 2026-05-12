"""
============================================================================
SCRAPER SMV vF_2025 - Descarga PDFs para Power Automate + AI Builder
============================================================================
Solo descarga y organiza. NO analiza contenido de PDFs.
Power Automate + AI Builder/Copilot extrae Auditor y Fee de auditoria.

  - Limpia PDFs y Excel anterior al iniciar (data fresca cada ejecucion)
  - Descarga TODOS los PDFs por empresa desde la SMV
  - Organiza en carpetas: {SharePoint}/pdfs/{Empresa}/
  - Excel con nombre fijo (smv_auditores.xlsx) para Power Automate
  - Tabla 'tblAuditores' creada automaticamente en el Excel
  - Columna 'Fecha Extraccion' con timestamp de ejecucion
  - Columnas vacias: Auditor y Fee Auditoria (Power Automate las llena)
  - Reintentos automaticos en descargas fallidas (3 intentos)

Flujo:
  1. Ejecutar este script -> limpia anterior + descarga nuevos + Excel
  2. OneDrive sincroniza automaticamente a SharePoint
  3. Ejecutar flujo de Power Automate (manual)

Requisitos:
    pip install selenium openpyxl webdriver-manager requests

Uso:
    python scraper_smv_vF_2025.py              # Completo (~25 min)
    python scraper_smv_vF_2025.py --max 10     # Prueba con 10 empresas
    python scraper_smv_vF_2025.py --no-headless
============================================================================
"""

import time, random, re, sys, argparse, logging, shutil
from datetime import datetime
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException

import requests as req
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================================
# CONFIGURACION
# ============================================================================
URL = (
    "https://www.smv.gob.pe/SIMV/Frm_InformacionFinanciera"
    "?data=A70181B60967D74090DCD93C4920AA1D769614EC12"
)
ANIOS = ["2025"]
DELAY = (2.0, 4.0)
MAX_REINTENTOS = 3
MIN_PDF_SIZE = 1024              # 1 KB minimo para considerar PDF valido
EXCEL_NAME = "smv_auditores.xlsx"  # Nombre fijo para que Power Automate siempre apunte al mismo archivo

# Ruta SharePoint sincronizada (OneDrive) - aqui va todo el output
#SHAREPOINT_DIR = Path(r"C:\Users\avillanuev044\OneDrive - PwC\PE-IFS-SALES - Entidades de Regulación\SMV")
SHAREPOINT_DIR = Path("output")
OUTPUT_DIR = SHAREPOINT_DIR                    # Excel va aqui
PDF_DIR = SHAREPOINT_DIR / "pdfs_smv"              # PDFs organizados por empresa
LOG_DIR = Path("output")                       # Logs se quedan local (no llenar SharePoint)
TS = datetime.now().strftime("%Y%m%d_%H%M%S")

# ============================================================================
# LOGGING
# ============================================================================
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
PDF_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_DIR / f"scraper_v11_{TS}.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# Session reutiliza conexiones TCP (mas rapido que req.get individual)
SESSION = req.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
})


# ============================================================================
# UTILIDADES
# ============================================================================
def limpiar_nombre_carpeta(nombre: str) -> str:
    """Limpia nombre de empresa para usarlo como nombre de carpeta.
    Remueve caracteres que causan error 404 en el conector SharePoint de Power Automate.
    """
    # Remover caracteres invalidos en Windows + problematicos en SharePoint API
    nombre = re.sub(r'[<>:"/\\|?*(){}[\]#%&!]', '', nombre)
    # Remover contenido entre parentesis ANTES de limpiar (ej: "(ANTES FALABELLA...)")
    # ya se removieron los parentesis arriba, pero el contenido queda
    nombre = re.sub(r'\s+', ' ', nombre.strip())
    # Remover puntos al final (SharePoint no permite carpetas terminando en punto)
    nombre = nombre.rstrip('. ')
    return nombre[:80].strip()


def extraer_tipo_sociedad(nombre: str) -> str:
    """Extrae el tipo de sociedad del nombre de la empresa.
    Ej: 'ALICORP S.A.A.' -> 'S.A.A.'
    """
    patrones = [
        (r'\bS\.?\s*A\.?\s*A\.?\b', 'S.A.A.'),          # Sociedad Anonima Abierta
        (r'\bS\.?\s*A\.?\s*C\.?\b', 'S.A.C.'),          # Sociedad Anonima Cerrada
        (r'\bS\.?\s*A\.?\b(?!\.\s*[AC])', 'S.A.'),      # Sociedad Anonima
        (r'\bS\.?\s*Civil\b', 'S. Civil'),               # Sociedad Civil
        (r'\bS\.?\s*C\.?\s*R\.?\s*L\.?\b', 'S.C.R.L.'), # Sociedad Civil Resp. Limitada
        (r'\bS\.?\s*R\.?\s*L\.?\b', 'S.R.L.'),          # Sociedad Resp. Limitada
        (r'\bSUCURSAL\b', 'Sucursal'),                   # Sucursal
    ]
    for patron, tipo in patrones:
        if re.search(patron, nombre, re.IGNORECASE):
            return tipo
    return ""


def limpiar_ejecucion_anterior():
    """
    Elimina PDFs de ejecuciones anteriores para garantizar data fresca.
    El Excel NO se elimina — se sobreescribe para que SharePoint mantenga
    el mismo ID interno y Power Automate no pierda la referencia.
    """

    # Limpiar carpeta de PDFs
    if not PDF_DIR.exists():
        PDF_DIR.mkdir(parents=True, exist_ok=True)
        log.info("Carpeta PDFs creada (no habia anterior)")
        return

    contenido = list(PDF_DIR.iterdir())
    if not contenido:
        log.info("Carpeta PDFs ya esta vacia, nada que limpiar")
        return

    total_pdfs = sum(1 for p in PDF_DIR.rglob("*.pdf"))
    total_carpetas = sum(1 for p in PDF_DIR.iterdir() if p.is_dir())

    log.info(f"Limpiando ejecucion anterior: {total_pdfs} PDFs en {total_carpetas} carpetas...")
    for intento in range(1, 4):
        try:
            shutil.rmtree(PDF_DIR)
            PDF_DIR.mkdir(parents=True, exist_ok=True)
            log.info("Limpieza completada. Listo para nueva descarga")
            break
        except PermissionError:
            if intento < 3:
                log.warning(f"  OneDrive bloqueando archivos, reintento {intento}/3 en 5s...")
                time.sleep(5)
            else:
                log.warning("  No se pudo limpiar carpeta PDFs (OneDrive sync). Continuando con carpeta existente...")
                PDF_DIR.mkdir(parents=True, exist_ok=True)


def crear_driver(headless=True):
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    try:
    #Operador Google
        from webdriver_manager.chrome import ChromeDriverManager
        svc = ChromeService(ChromeDriverManager().install())
    except ImportError:
        svc = ChromeService()
    #svc = ChromeService(r"C:\Users\avillanuev044\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe")
    d = webdriver.Chrome(service=svc, options=opts)
    d.set_page_load_timeout(45)
    d.implicitly_wait(8)
    return d


def descargar_pdf(url: str, ruta_destino: Path) -> bool:
    """
    Descarga un PDF y lo guarda en disco.
    - Reintentos automaticos (3 intentos con backoff)
    - Solo guarda PDFs reales (valida magic bytes %PDF-)
    - Valida tamano minimo (>1 KB)
    """
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            r = SESSION.get(url, timeout=45, allow_redirects=True)
            if r.status_code != 200:
                log.warning(f"    HTTP {r.status_code} para {ruta_destino.name} (intento {intento})")
                continue

            data = r.content

            # Solo guardar si es PDF real (magic bytes %PDF-)
            if data[:5] != b'%PDF-':
                log.warning(f"    {ruta_destino.name}: no es PDF (posible pagina de error HTML)")
                return False

            # Validar tamano minimo
            if len(data) < MIN_PDF_SIZE:
                log.warning(f"    {ruta_destino.name}: PDF muy pequeno ({len(data)} bytes), descartado")
                return False

            ruta_destino.write_bytes(data)
            log.info(f"    Guardado: {ruta_destino.name} ({len(data)//1024} KB)")
            return True

        except req.exceptions.Timeout:
            log.warning(f"    Timeout {ruta_destino.name} (intento {intento}/{MAX_REINTENTOS})")
        except req.exceptions.ConnectionError:
            log.warning(f"    Error conexion {ruta_destino.name} (intento {intento}/{MAX_REINTENTOS})")
        except Exception as e:
            log.warning(f"    Error descarga {ruta_destino.name}: {e} (intento {intento}/{MAX_REINTENTOS})")

        if intento < MAX_REINTENTOS:
            time.sleep(3 * intento)  # Backoff: 3s, 6s

    log.error(f"    FALLO definitivo: {ruta_destino.name} despues de {MAX_REINTENTOS} intentos")
    return False


def limpiar_carpetas_vacias():
    """Elimina subcarpetas vacias al final para no confundir a Power Automate.
    Ignora errores de permisos (OneDrive puede bloquear carpetas durante sync).
    """
    eliminadas = 0
    for carpeta in PDF_DIR.iterdir():
        if carpeta.is_dir() and not any(carpeta.iterdir()):
            try:
                carpeta.rmdir()
                eliminadas += 1
            except PermissionError:
                log.warning(f"  No se pudo eliminar carpeta (OneDrive sync): {carpeta.name}")
            except Exception as e:
                log.warning(f"  Error eliminando carpeta {carpeta.name}: {e}")
    if eliminadas:
        log.info(f"  Carpetas vacias eliminadas: {eliminadas}")


# ============================================================================
# SCRAPING PRINCIPAL
# ============================================================================
def scrape(driver, max_emp=None):
    driver.get(URL)
    time.sleep(5)
    sel = Select(driver.find_element(By.ID, "MainContent_cboDenominacionSocial"))
    empresas = [(o.get_attribute("value"), o.text.strip()) for o in sel.options
                if o.get_attribute("value") and o.text.strip()
                and "ingrese" not in o.text.lower()]
    log.info(f"Empresas encontradas: {len(empresas)}")
    if max_emp:
        empresas = empresas[:max_emp]
        log.info(f"Limitado a {max_emp}")

    resultados = []
    for idx, (val, nombre) in enumerate(empresas):
        log.info(f"[{idx+1}/{len(empresas)}] {nombre}")

        nombre_carpeta = limpiar_nombre_carpeta(nombre)
        carpeta_empresa = PDF_DIR / nombre_carpeta
        carpeta_empresa.mkdir(parents=True, exist_ok=True)

        reg = {
            "razon_social": nombre,
            "ruc": val if val.isdigit() and len(val) == 11 else "",  # RUC si el value del dropdown es de 11 digitos
            "tipo_sociedad": extraer_tipo_sociedad(nombre),
            "carpeta": nombre_carpeta,
            "auditor": "",                  # Power Automate llena esto
            "fee": "",                      # Power Automate llena esto (honorarios)
            "tipo_opinion": "",             # Power Automate llena esto
            "directorio": "",               # Power Automate llena esto
            "estado": "",                   # OK / Sin dictamen / Error
            "nro_expediente": "",
            "fecha_presentacion": "",
            "periodo": "2025",
            "pdfs_descargados": 0,
            "pdfs_nombres": "",
        }

        # Retry para errores de red (VPN/DNS)
        max_intentos_red = 3
        for intento_red in range(1, max_intentos_red + 1):
            try:
                driver.get(URL)
                time.sleep(2.5)
                Select(driver.find_element(
                    By.ID, "MainContent_cboDenominacionSocial"
                )).select_by_value(val)
                time.sleep(1)

                radio = driver.find_element(By.ID, "MainContent_cboPeriodo_1")
                if not radio.is_selected():
                    radio.click()
                    time.sleep(2)

                anio_encontrado = False
                for anio in ANIOS:
                    try:
                        Select(driver.find_element(
                            By.ID, "MainContent_cboAnio"
                        )).select_by_value(anio)
                        reg["periodo"] = anio
                        anio_encontrado = True
                        break
                    except Exception:
                        continue
                time.sleep(0.5)
                # --- NUEVA REGLA ESTRICTA ---
                # Si no pudo seleccionar el año, aborta la búsqueda para esta empresa
                if not anio_encontrado:
                    log.info(f"  Año {ANIOS[0]} no disponible en la SMV.")
                    reg["estado"] = "Sin dictamen"
                    break # Sale del intento y pasa a la siguiente empresa de la lista

                driver.find_element(By.ID, "MainContent_cbBuscar").click()
                time.sleep(4)

                # Intentar extraer RUC de la pagina (buscar en el HTML)
                if not reg["ruc"]:
                    try:
                        page_text = driver.page_source
                        ruc_match = re.search(r'\b(20\d{9})\b', page_text)
                        if ruc_match:
                            reg["ruc"] = ruc_match.group(1)
                    except Exception:
                        pass

                try:
                    tabla = driver.find_element(By.ID, "MainContent_grdInfoFinanciera")
                except NoSuchElementException:
                    log.info("  Sin resultados en SMV")
                    reg["estado"] = "Sin dictamen"
                    break

                # Recorrer TODAS las filas, descargar TODOS los PDFs
                pdfs_guardados = []

                for fila in tabla.find_elements(By.TAG_NAME, "tr")[1:]:
                    celdas = fila.find_elements(By.TAG_NAME, "td")
                    if len(celdas) < 4:
                        continue

                    tipo_doc = celdas[1].text.strip()
                    nro_exp = celdas[2].text.strip()
                    fecha = celdas[3].text.strip()

                    # Saltar archivos XBRL (son HTML, no PDF - ahorra ~114 descargas)
                    if "xbrl" in tipo_doc.lower() or "archivo estructurado" in tipo_doc.lower():
                        continue

                    if not reg["nro_expediente"]:
                        reg["nro_expediente"] = nro_exp
                        reg["fecha_presentacion"] = fecha

                    for lk in celdas[-1].find_elements(By.TAG_NAME, "a"):
                        href = lk.get_attribute("href") or ""
                        if "ConsultasP8" not in href and "documento" not in href.lower():
                            continue

                        tipo_limpio = re.sub(r'[<>:"/\\|?*]', '', tipo_doc)
                        tipo_limpio = re.sub(r'\s+', '_', tipo_limpio.strip())[:50]
                        pdf_nombre = f"{tipo_limpio}_{reg['periodo']}.pdf"

                        ruta_pdf = carpeta_empresa / pdf_nombre
                        contador = 1
                        while ruta_pdf.exists():
                            contador += 1
                            pdf_nombre = f"{tipo_limpio}_{reg['periodo']}_{contador}.pdf"
                            ruta_pdf = carpeta_empresa / pdf_nombre

                        if descargar_pdf(href, ruta_pdf):
                            pdfs_guardados.append(pdf_nombre)

                reg["pdfs_descargados"] = len(pdfs_guardados)
                reg["pdfs_nombres"] = " | ".join(pdfs_guardados)
                reg["estado"] = "OK" if pdfs_guardados else "Sin dictamen"

                if pdfs_guardados:
                    log.info(f"  {len(pdfs_guardados)} PDFs descargados")
                else:
                    log.info("  Sin PDFs disponibles")

                break  # Exito — sale del retry

            except Exception as e:
                error_msg = str(e)
                if "ERR_NAME_NOT_RESOLVED" in error_msg or "net::" in error_msg or "Timed out" in error_msg:
                    if intento_red < max_intentos_red:
                        log.warning(f"  Error de red (intento {intento_red}/{max_intentos_red}): {error_msg[:60]}. Reintentando en 10s...")
                        time.sleep(10)
                        continue
                    else:
                        log.error(f"  Error de red persistente despues de {max_intentos_red} intentos: {error_msg[:80]}")
                        reg["estado"] = "Error"
                else:
                    log.error(f"  Error: {error_msg[:120]}")
                    reg["estado"] = "Error"
                break

        resultados.append(reg)
        time.sleep(random.uniform(*DELAY))

    # Limpiar carpetas vacias al final (no critico - no debe bloquear el Excel)
    try:
        limpiar_carpetas_vacias()
    except Exception as e:
        log.warning(f"  No se pudieron limpiar carpetas vacias: {e}")

    return resultados


# ============================================================================
# EXPORTAR EXCEL
# ============================================================================
def exportar(datos, filepath):
    wb = Workbook()
    HF = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HFL = PatternFill("solid", fgColor="1F4E79")
    HA = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CF = Font(name="Arial", size=10)
    CA = Alignment(vertical="top", wrap_text=True)
    BRD = Border(
        left=Side("thin", "B0B0B0"), right=Side("thin", "B0B0B0"),
        top=Side("thin", "B0B0B0"), bottom=Side("thin", "B0B0B0"),
    )
    ALT = PatternFill("solid", fgColor="E8F0FE")
    ERR_FILL = PatternFill("solid", fgColor="FCE4EC")

    # --- Hoja principal: para Power Automate ---
    ws = wb.active
    ws.title = "Empresas_Auditores"
    headers = [
        ("Razon Social", 55),
        ("RUC", 16),                      # Identificador unico
        ("Tipo Sociedad", 14),            # S.A.A., S.A.C., S.A., etc.
        ("Carpeta", 40),
        ("Auditor", 50),              # Power Automate escribe aqui
        ("Fee Auditoria", 18),        # Power Automate escribe aqui (honorarios)
        ("Tipo de Opinion", 22),      # Power Automate escribe aqui
        ("Directorio", 50),           # Power Automate escribe aqui
        ("Fuente", 10),               # SMV para todos
        ("Estado", 14),               # OK / Sin dictamen / Error
        ("Nro. Expediente", 16),
        ("Periodo", 10),
        ("Fecha Presentacion", 20),
        ("Fecha Extraccion", 20),     # Cuando se ejecuto el script
        ("PDFs Descargados", 10),
        ("Nombres PDFs", 60),
    ]
    for i, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = HF, HFL, HA, BRD
        ws.column_dimensions[get_column_letter(i)].width = w

    ds = sorted(datos, key=lambda x: x["razon_social"].upper())
    fecha_extraccion = datetime.now().strftime("%d/%m/%Y %H:%M")

    for i, r in enumerate(ds, 2):
        es_error = r["estado"] in ("Error", "Sin dictamen")
        fill = ERR_FILL if es_error else (ALT if i % 2 == 0 else PatternFill())
        values = [
            r["razon_social"],
            r["ruc"],
            r["tipo_sociedad"],
            r["carpeta"],
            r["auditor"],
            r["fee"],                 # Vacio - Power Automate llena
            r["tipo_opinion"],        # Vacio - Power Automate llena
            r["directorio"],          # Vacio - Power Automate llena
            "SMV",                    # Fuente fija
            r["estado"],
            r["nro_expediente"],
            r["periodo"],
            r["fecha_presentacion"],
            fecha_extraccion,         # Misma fecha para todas las filas
            r["pdfs_descargados"],
            r["pdfs_nombres"],
        ]
        for col, v in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font, c.alignment, c.border, c.fill = CF, CA, BRD, fill

    # Crear tabla tblAuditores para Power Automate
    last_row = len(ds) + 1
    last_col = get_column_letter(len(headers))  # P (16 columnas)
    table_ref = f"A1:{last_col}{last_row}"
    tabla = Table(displayName="tblAuditores", ref=table_ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tabla)
    ws.freeze_panes = "A2"

    # --- Hoja 2: Resumen ---
    ws2 = wb.create_sheet("Resumen")
    total = len(ds)
    ok = sum(1 for d in ds if d["estado"] == "OK")
    sin_dict = sum(1 for d in ds if d["estado"] == "Sin dictamen")
    errores = sum(1 for d in ds if d["estado"] == "Error")
    total_pdfs = sum(d["pdfs_descargados"] for d in ds)

    metricas = [
        ("Total empresas", total),
        ("Con PDFs descargados (OK)", ok),
        ("Sin dictamen en SMV", sin_dict),
        ("Errores de scraping", errores),
        ("Total PDFs descargados", total_pdfs),
        ("", ""),
        ("Carpeta PDFs", str(PDF_DIR.resolve())),
        ("", ""),
        ("Paso siguiente", "Ejecutar flujo Power Automate"),
        ("", "(Excel y PDFs ya estan en SharePoint via OneDrive)"),
    ]
    for i, (k, v) in enumerate(metricas, 1):
        ws2.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=i, column=2, value=str(v)).font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 60

    wb.save(filepath)
    log.info(f"\nExcel guardado: {filepath}")
    log.info(f"  Total:           {total}")
    log.info(f"  OK (con PDFs):   {ok}")
    log.info(f"  Sin dictamen:    {sin_dict}")
    log.info(f"  Errores:         {errores}")
    log.info(f"  Total PDFs:      {total_pdfs}")
    log.info(f"  Carpeta PDFs:    {PDF_DIR.resolve()}")
    log.info(f"\n>> Excel y PDFs guardados en SharePoint (via OneDrive)")
    log.info(f">> Ejecuta el flujo de Power Automate para extraer auditores")


# ============================================================================
# MAIN
# ============================================================================
def main():
    p = argparse.ArgumentParser(description="Scraper SMV v11 - Descarga para Power Automate")
    p.add_argument("--no-headless", action="store_true",
                   help="Mostrar navegador")
    p.add_argument("--max", type=int, default=None,
                   help="Limitar cantidad de empresas (para pruebas)")
    args = p.parse_args()

    log.info("=" * 60)
    log.info("SCRAPER SMV v11 - Descarga PDFs para Power Automate")
    log.info("=" * 60)

    # Limpiar ejecucion anterior (PDFs + Excel viejos)
    limpiar_ejecucion_anterior()

    driver = None
    try:
        driver = crear_driver(headless=not args.no_headless)
        log.info("Chrome OK")
        datos = scrape(driver, args.max)
        if datos:
            excel_path = OUTPUT_DIR / EXCEL_NAME
            exportar(datos, excel_path)
        try:
            driver.save_screenshot(str(LOG_DIR / f"screenshot_v11_{TS}.png"))
        except Exception:
            pass
    except Exception as e:
        log.error(f"Error fatal: {e}")
    finally:
        if driver:
            driver.quit()
        SESSION.close()


if __name__ == "__main__":
    main()
